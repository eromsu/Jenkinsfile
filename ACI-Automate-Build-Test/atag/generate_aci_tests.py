import argparse
import os
import jinja2
import openpyxl
import warnings
import sys
import time
import fnmatch
from jinja2 import Environment, FileSystemLoader
import aci_input_validation as validate
import toolUsage as usageInfo
import yaml
from builtins import input
import pkg_resources
import re

### Check OpenpyXL Version and print warning if version post 2.4
### To Do adapt openpyxl function call based on version to avoid deprecation warning message
version_match = re.compile(r'^2\.4')
openpyxl_version = pkg_resources.get_distribution("openpyxl").version
if not re.match(version_match,openpyxl_version):
    print ("WARNING !! : This script best works with OPENPYXL version 2.4.8 you are currenyl using version: %s\n" % openpyxl_version)
    OPENPYXLMODE = "2.5"
else:
    OPENPYXLMODE = "2.4"

warnings.filterwarnings("ignore")  # Disable warning message logged by openpyxl

# BAECKS - 2018/06/28: The code below is required to have YAML parsing and dumping maintain the order of the parameters.
# The order has no importance in a YAML file, but
from collections import OrderedDict

try:
    from yaml import CLoader as Loader, CDumper as Dumper
except ImportError:
    from yaml import Loader, Dumper
from yaml.representer import SafeRepresenter
_mapping_tag = yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG

def dict_representer(dumper, data):
    return dumper.represent_dict(data.items())

def dict_constructor(loader, node):
    return OrderedDict(loader.construct_pairs(node))

Dumper.add_representer(OrderedDict, dict_representer)
Loader.add_constructor(_mapping_tag, dict_constructor)

Dumper.add_representer(str,
                       SafeRepresenter.represent_str)
# BAECKS - 2018/06/28 ---

class DafeRow(object):
    def __init__(self, row_data):
        for parameter, value in row_data.items():
            if parameter == None:
                continue
            param = parameter.replace(" ", "_").lower().strip()
            self.__setattr__(param, value)


class DafeSheet(list):
    def __init__(self, sheet_data):
        for row in sheet_data:
            self.append(DafeRow(row))

    def row(self, **kwargs):
        for rw in self:
            match = True
            for param, value in kwargs.items():
                try:
                    actual_value = getattr(rw, param)
                    if actual_value.lower() != value.lower():
                        match = False
                        break
                except:
                    return None  # if the parameter does not exist, then there is no use in continuing next rows
            if match:
                return rw

        return None


class DafeConfig(object):
    def __init__(self, data_config):
        for sheet, data in data_config.items():
            self.__setattr__(sheet, DafeSheet(data))


class ExecException(Exception):
    """
    Exception Class
    """

    def fatal(self, msg):
        """
        Prints and error message and aborts program execution
        """
        sys.stderr.write(msg + "\n")
        sys.exit(1)

    def warning(self, msg):
        """
        Prints a warning message to stderr
        """
        sys.stderr.write(msg + "\n")


class OutputLogging():
    """
    Output Logging
    Used to print to STDOUT and or LOGFILE
    Print to stdout will be enhanced in order to pretty print the messages
    """

    def __init__(self, method, file_name):
        now = time.time()
        local_time = time.localtime(now)
        timestamp = time.strftime("%Y-%m-%d-%H%M", local_time)

        # Validate method input
        assert method.lower() == "console-only" or method.lower() == "file", "Please use either console-only or" \
                                                                             " file as logging method"

        self.method = method.lower()
        if self.method == "file":
            self.output_file = timestamp + "-" + file_name
            self.output = open(self.output_file, "a")

    def print_string_to_file(self, message):
        self.output.write(message + "\n")

    def print_to_stdout(self, message):
        print(message)

    def print_message(self, message):
        """ Print message to both STDOUT and output Logfile"""
        self.print_to_stdout(message)
        if self.method == "file":
            self.print_string_to_file(message)


LOG = OutputLogging("console-only", "")
ExecERROR = ExecException()


class TemplateRender():
    """ Template rendering Class which use jinja2 Templating language
    The class constructor takes the path to the xml template directory as input data
    The file names in the directory are loaded in a dictionary structure
    The class provides the  method render_template for rendering a template
    """

    def __init__(self, template_path):
        self.jinja_env = self.create_jinja_env(template_path)
        self.template_dict = self.load_template(template_path)

    def create_jinja_env(self, template_path):
        template_loader = FileSystemLoader(searchpath=template_path)
        template_env = Environment(loader=template_loader, trim_blocks=True, lstrip_blocks=True)
        return template_env

    def load_template(self, template_path):
        try:
            template_dict = {}
            for template in self.files_within(template_path):
                filename = template.split(template_path)
                relative_path = filename[1]
                filename = relative_path.split(".")
                fo = open(template, "r")
                template_dict[filename[0]] = fo.read()
            return template_dict
        except OSError as e:
            ExecERROR.fatal("ERROR: Can't find template directory %s " % template_path)
            sys.exit(1)
        except:
            ExecERROR.fatal("ERROR: Undefined error in loading template")

    def files_within(self, directory_path, pattern="*"):
        for dirpath, dirnames, filenames in os.walk(directory_path):
            for file_name in fnmatch.filter(filenames, pattern):
                yield os.path.join(dirpath, file_name)

    def render_template(self, template_file, item, full_config, **kwargs):
        """ Takes as argument
        The name of the template file to be used : template_file
        """
        try:
            template = self.jinja_env.get_template(template_file)
            xml_data = template.render(config=item, dafe_data=full_config, **kwargs)
            return xml_data
        except jinja2.TemplateNotFound as e:
            ExecERROR.fatal("ERROR: can't find template file %s" % template_file)
        except jinja2.TemplateSyntaxError as e:
            message = "ERROR: Syntax Error in Template %s " % template_file
            error_line = str(e.message) + " line number : " + str(e.lineno)
            ExecERROR.warning(message + " " + error_line)
        except Exception as e:
            ExecERROR.warning(" Undefined error while rendering template %s" % template_file)


class InputData(object):
    """
    This class provides methods to manipulate data input
    Create the build task
    Load excel workbook and convert into dictionary
    Call aci_input_validation module to validate data
    """

    def open_excel_workbook(self, file_name):
        """ Takes file name as input and return a openpyxl workbook object
        @param file_name : excel source file name
        @return workbook : openpyxl workbook object
        """
        try:
            LOG.print_message("Opening excel workbook %s" % file_name)
            workbook = openpyxl.load_workbook(file_name,
                                              data_only=True)  # data_only = True is used to load cell data value only not formula
            return workbook
        except IOError as e:
            ExecERROR.fatal("IOError Can't Open file %s" % file_name)
        except:
            ExecERROR.fatal("Undefined error opening excel %s data input" % file_name)

    def excel_to_dictionary(self, workbook, build_tasks):
        """ Takes a workbook and a worksheet name as input
        return a dictionary with the following structure
        {sheet_name:[{},{},...]}
        """
        active_dict = {}
        LOG.print_message("+---------- Loading Input Data ----------+")
        # for task in build_tasks:
        #    sheet_name = task[1]
        for sheet_name in workbook.get_sheet_names():
            # Handle undefined sheet name
            if sheet_name == '':
                active_dict[sheet_name] = []
                new_entry = {}
                active_dict[sheet_name].append(new_entry)
            # Skip import of 'Fabric Configuration Steps' sheet, as this one does not contain date that are needed later on
            elif sheet_name == 'Fabric Configuration Steps':
                continue
            else:
                LOG.print_message("+--- Importing worksheet %s" % sheet_name)
                if OPENPYXLMODE == "2.5":
                    active_sheet = workbook[sheet_name]
                else:
                    active_sheet = workbook.get_sheet_by_name(sheet_name)
                active_dict[sheet_name] = []
                header = []
                for col_num in range(1, active_sheet.max_column + 1):
                    header.append(active_sheet.cell(row=1, column=col_num).value)
                for row_num in range(2, active_sheet.max_row + 1):
                    new_entry = {}
                    for col_num in range(1, active_sheet.max_column + 1):
                        if active_sheet.cell(row=row_num, column=col_num).value != None:
                            new_entry[header[col_num - 1]] = str(
                                active_sheet.cell(row=row_num, column=col_num).value).lstrip().rstrip()
                        else:
                            new_entry[header[col_num - 1]] = ""
                    active_dict[sheet_name].append(new_entry)
        LOG.print_message("+------------------------------------------+")
        return active_dict

    def validate_input(self, build_tasks, datadict):
        """
        """
        is_data_valid = True
        LOG.print_message("+---------- Input Data Validation ----------+")
        with open("data_validation.yaml", "r") as f:
            validation_scheme = yaml.load(f)
        for task in build_tasks:    
            object_type, worksheet_name, template_file_name, test_type, test_variables = task[:5]
            # Loop through each test type (configuration, faults, etc.)
            for test in test_type:

                # Only validate inputs based on a worksheet
                try:
                    if datadict[worksheet_name]:
                        if validation_scheme[test].has_key(template_file_name):
                            LOG.print_message("+- Validating sheet '%s' for template '%s' (%s)" % (worksheet_name, template_file_name, test))
                            validation_result = validate.validate_input(datadict[worksheet_name],
                                                                        validation_scheme[test][template_file_name])
                            #validation_result = validate.validate_input(datadict[worksheet_name],
                            #                                            validation_scheme[template_file_name])
                            if len(validation_result) != 0:
                                LOG.print_message("+-       Incorrect input in sheet '%s'   Line %s" % (worksheet_name, validation_result))
                                is_data_valid = False
                        else:
                            LOG.print_message("+- Skipping validation for sheet '%s' no validation scheme for template '%s' (%s)" % (
                            worksheet_name, template_file_name, test))
                except KeyError:
                    if validation_scheme[test].has_key(template_file_name):
                        LOG.print_message("+- Skipping validation (no input sheet used) for template '%s' (%s)" % (
                                template_file_name, test))
                    else:
                         LOG.print_message("+- Skipping validation (no input sheet used) no validation scheme for template '%s' (%s)" % (
                                template_file_name, test))
                                      
        LOG.print_message("+------------------------------------------+")
        if not is_data_valid:
            ExecERROR.fatal("Found incorrect input data, please fix before proceeding ")


class InputConfig():
    """
    This class provides methods to manipulate input configuration data
    Create the build task
    Merge configuration with dictionary from excel workbook
    """

    def __init__(self, config_file):
        self.config = self.read_config(config_file)

    def read_config(self, cfg_file):
        """
        """
        try:
            with open(cfg_file, 'r') as ymlfile:
                cfg_dict = yaml.load(ymlfile)
            return cfg_dict
        except:
            ExecERROR.fatal("ERROR: Error reading config_file %s" % cfg_file)
            sys.exit(1)

    def get_template_dir(self):
        return (self.config['testcase_config']['template_directory'])

    def get_input_file(self):
        return (self.config['testcase_config']['dafe_workbook'])

    def get_rasta_robot_header(self):
        return (self.config['testcase_templates']['rasta_robot_header']['template'])

    def get_testbed_file(self):
        try:
            return self.config['testcase_config']['testbed_file']
        except:
            return None

    def get_device_username(self):
        try:
            return self.config['testcase_config']['device_username']
        except:
            return None

    def get_device_password(self):
        try:
            return self.config['testcase_config']['device_password']
        except:
            return None

    def get_build_tasks(self):
        """
        """
        # Tenerate default test type list
        default_test_type = self.config['testcase_config']['default_test_type']
        default_fault_variables = self.config['testcase_config']['default_faults_variables']

        LOG.print_message("+---------- Creating Build tasks ----------+")
        task_lists = []
        tests = self.config['tests_to_generate']
        for test_category in tests:
            test_category_name = test_category['category']
            # Generate list of test_case types
            if not 'test_type' in test_category:
                test_category['test_type'] = default_test_type[:]

            # Skip test_type(s) where the associated template does not exist
            test_type_candidates = test_category['test_type'][:]
            for test_type in test_category['test_type']:
                template_file_path = self.config['testcase_config']['template_directory'] + test_type + "/" + self.config['testcase_templates'][test_category_name]['template']
                if not os.path.isfile(template_file_path):
                    test_type_candidates.remove(test_type)
            test_category['test_type'] = test_type_candidates[:]

            # Generate custom variables
            test_variables = {}
            for test_type in test_category['test_type']:
                variable_name = test_type + '_variables'
                if not variable_name in test_category:
                    if test_type == "faults":
                        test_variables[test_type] = default_fault_variables
                    else:
                        test_variables[test_type] = []
                else:
                    test_variables[test_type] = test_category[variable_name]

            # Create build tasks
            if 'test_enabled' in test_category:
                if test_category['test_enabled'] == 'yes':
                    new_entry = []
                    new_entry.append(self.config['testcase_templates'][test_category_name]['description'])
                    new_entry.append(self.config['testcase_templates'][test_category_name]['input_sheet'])
                    new_entry.append(self.config['testcase_templates'][test_category_name]['template'])
                    new_entry.append(test_category['test_type'])
                    new_entry.append(test_variables)
                    task_lists.append(new_entry)
                    LOG.print_message("+ " + new_entry[0])

        LOG.print_message("+------------------------------------------+")
        return task_lists

# BAECKS - 2018/06/28
class Testbed():
    """
    This class allows reading in a base testbed file (YAML format), updating the content (e.g. adding devices)
    and outputting this content in YAML format.
    """
    def __init__(self, input_file, config_data):
        self.config_data = config_data

        if input_file == None:
            self.testbed = []
        else:
            try:
                with open(input_file, 'r') as ymlfile:
                    self.testbed = yaml.load(ymlfile)
            except:
                ExecERROR.fatal("ERROR: Error reading input testbed file %s" % input_file)
                sys.exit(1)

    def update_devices(self, device_username, device_password, test_template):
        if device_username == None:
            return

        if device_password == None:
            return

        testbed_yaml = test_template.render_template("testbed_template.yaml", None, self.config_data,
                                                     username=device_username, password=device_password)
        testbed_data = yaml.load(testbed_yaml)
        try:
            self.testbed['devices'].update(testbed_data['devices'])
        except:
            self.testbed['devices'] = testbed_data['devices']

    def write_to_file(self, file_name):
        """
        Dumps the testbed data to a YAML file that can be used in RASTA tests.
        :param file_name: The path for the output file
        :return:
        """
        try:
            with  open(file_name, "w") as testbed_file:
                yaml.dump(self.testbed, testbed_file, default_flow_style=False)
        except Exception as e:
            ExecERROR.fatal("ERROR: Error writing testbed file %s (%s)" % (file_name, str(e)))
            sys.exit(1)
# BAECKS - 2018/06/28 ---

class TestcaseBuild():
    """ Takes as argument
    The path to the configuration file folder : config_file
    The name of the output data file : output_file
    """

    def __init__(self, config_file, output_file, password, force_continue):
        self.config = InputConfig(config_file)
        self.password = password
        self.username = None
        self.test_template = TemplateRender(self.config.get_template_dir())
        self.tool_usage = {}
        self.tool_usage["mode"] = "offline"
        self.tool_stats = {}

        self.build_tasks = self.config.get_build_tasks()

        self.input_data = InputData()
        self.output_file = output_file

        if not force_continue:
            if input(" Do you want to proceed with the above testcase tasks y/n ? [n] : ") != 'y':
                LOG.print_message("Aborting configuration")
                return
            else:
                LOG.print_message("Proceeding....")

        self.workbook = self.input_data.open_excel_workbook(self.config.get_input_file())
        self.generate_testcases(self.build_tasks)

    def get_state_mode(self, data_dictionary):
        """ Planned for future feature   deleting ignoring specific lines in the excel worksheet
        """
        try:
            mode = data_dictionary['status']
            if mode == "deleted":
                action = "deleting"
            elif mode == "ignored":
                action = "skipping"
            else:
                action = "creating"
            return action
        except:
            action = "creating"
            return action

    def generate_testcases(self, build_tasks):
        test_data = ""
        datadict = self.input_data.excel_to_dictionary(self.workbook, build_tasks)
        self.input_data.validate_input(build_tasks, datadict)

        full_config = DafeConfig(datadict)

        testbed_file = os.path.splitext(self.output_file)[0] + "_testbed.yaml"

        testcase_header = self.test_template.render_template(self.config.get_rasta_robot_header(), "", full_config,
                                                             testbed_file=os.path.basename(testbed_file)) + "\n"

        # Initiate tool usage infomation
        self.tool_usage["build_tasks"] = []
        self.tool_stats["summary_count"] = 0
        self.tool_stats["configuration"] = 0
        self.tool_stats["faults"] = 0
        self.tool_stats["operational_state"] = 0

        LOG.print_message("+------- Creating RASTA/CXTA Test-cases --------+")
        for task in build_tasks:
            # Record tool usage
            task_usage = []
            test_usage = []
            task_usage.append(task[1])
            object_type, worksheet_name, template_file_name, test_type, test_variables = task[:5]
            # Loop through each test type (configuration, faults, etc.)
            for test in test_type:
                # Record tool usage
                test_usage.append(test)
                if not test in self.tool_stats.keys():
                    self.tool_stats[test] = 0

                template_name = test + "/" + template_file_name

                if worksheet_name.strip() == '':
                    message = "+--- Generating %s test-cases (%s)" % (object_type, test)
                    LOG.print_message(message)

                    # Record tool usage
                    self.tool_stats[test] = self.tool_stats[test] + 1
                    self.tool_stats["summary_count"] = self.tool_stats["summary_count"] + 1

                    # Add test type specific variables
                    item = {}
                    if len(test_variables[test]) > 0:
                        for key in test_variables[test]:
                            item[key] = test_variables[test][key]

                    test_data = test_data + self.test_template.render_template(template_name, item, full_config) + "\n"
                    continue
                else:
                    try:
                        sheet = datadict[worksheet_name]
                    except:
                        LOG.print_message("The sheet %s doesn't exist! Skipping tests %s!" % (worksheet_name, test))
                        continue

                    try:
                        message = "+--- Generating %s test-cases (%s)" % (object_type, test)
                        LOG.print_message(message)
                        for item in datadict[worksheet_name]:
                            # Add test type specific variables
                            if len(test_variables[test]) > 0:
                                for key in test_variables[test]:
                                    item[key] = test_variables[test][key]

                            action = self.get_state_mode(item)
                            if action == "skipping" or action == "deleting":
                                try:
                                    LOG.print_message("+--- Skipping %s %s" % (object_type, item['name']))
                                except:
                                    excel_line_offset = 2
                                    LOG.print_message("+--- Skipping %s line %s" % (object_type, str(datadict[worksheet_name].index(item)+ excel_line_offset)))
                            else:
                                # Record tool usage
                                self.tool_stats[test] = self.tool_stats[test] + 1
                                self.tool_stats["summary_count"] = self.tool_stats["summary_count"] + 1

                                test_data = test_data + self.test_template.render_template(template_name, item,
                                                                                        full_config) + "\n"
                    except Exception as e:
                        ExecERROR.fatal("+--- Undefined error quitting ATAG")

            # Record tool usage
            task_usage.append(test_usage)
            self.tool_usage["build_tasks"].append(task_usage)

        output = open(self.output_file, "w")
        output.write(testcase_header)
        output.write(test_data)
        output.close()

        # BAECKS - 2018/06/28 --- testbed file
        testbed_input_file = self.config.get_testbed_file()
        testbed = Testbed(testbed_input_file, full_config)

        # If a username/password is provided on the command line, we use that one. If not we try obtaining it
        # from the config file. If no username/password is provided at all, no devices get added to the testbed file.
        device_username = self.username or self.config.get_device_username()
        device_password = self.password or self.config.get_device_password()
        if device_password != None and device_username != None:
            testbed.update_devices(device_username, device_password, self.test_template)

        testbed.write_to_file(testbed_file)

        # Present and submit tool usage information
        message = "+------------------------------------------------------+"
        LOG.print_message(message)
        message = "+---------- Test Case Generation Statistics -----------+"
        LOG.print_message(message)
        message = "+------------------------------------------------------+"
        LOG.print_message(message)
        message = "+- %s test cases generated in total" % self.tool_stats["summary_count"]
        LOG.print_message(message)
        message = "+--- %s configuration test cases " % self.tool_stats["configuration"]
        LOG.print_message(message)
        message = "+--- %s fault test cases " % self.tool_stats["faults"]
        LOG.print_message(message)
        message = "+--- %s operational state test cases " % self.tool_stats["operational_state"]
        LOG.print_message(message)

        # Submit tool usage information if script not executed in incognito mode
        if not cmd_parameters['incognito']:
            toolUsageUrl = "http://cx-emear-tools-stats.cisco.com/submit_tool_usage"
            submit_result = usageInfo.submit_usage_data("ATAG",toolUsageUrl, self.tool_usage, self.tool_stats)

            if submit_result["upload_status"] == "Success":
                message = "+- Usage statistics successfully submitted"
                LOG.print_message(message)
            else:
                message = "+- Failure during submission of usage statistics (can safely be ignorred)"
                LOG.print_message(message)

''' Grab INPUT_FILE from command arguments and perform basic validation '''
parser = argparse.ArgumentParser(
    description='''This scrips generates RASTA/CXTA .robot files based on a DAFE compatiple Excel workbook''')
parser.add_argument('-c', '--config', help='ATAG config file', dest="config",
                    default="atag_config.yaml", required=False)
parser.add_argument('-o', '--output', help='Output RASTA/CXTA test script file name', dest="output",
                    default="aci_tests.robot", required=False)
parser.add_argument('-p', '--password', help='password for APIC UI and device CLI access', dest="password",
                    default=None, required=False)
parser.add_argument('-f', '--force', action='store_const', const=True, default=False)
parser.add_argument('-i', '--incognito', action='store_const', const=True, default=False, help='Disables submission of tool usage information (NOT RECOMMENDED')
args = parser.parse_args()
cmd_parameters = vars(args)

TestcaseBuild(cmd_parameters['config'], cmd_parameters['output'], cmd_parameters['password'], cmd_parameters['force'])
